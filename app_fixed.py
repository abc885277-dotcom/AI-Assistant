import io
import json
import math
import os
import re
from typing import Any

import pandas as pd
import streamlit as st

# Optional imports are kept inside functions where practical so the UI can still
# explain a missing dependency instead of crashing at startup.
try:
    import ezdxf
except Exception:
    ezdxf = None

try:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter
except Exception:
    Workbook = None

try:
    from google import genai
except Exception:
    genai = None


st.set_page_config(
    page_title="AI BBS Engineer",
    page_icon="🏗️",
    layout="wide",
)

APP_TITLE = "🏗️ AI BBS Engineer"
APP_VERSION = "1.0.0"


def init_state() -> None:
    defaults = {
        "bars": [],
        "bbs_df": None,
        "analysis": "",
        "dxf_summary": None,
        "project_name": "BBS Project",
    }
    for key, value in defaults.items():
        if key not in st.session_state:
            st.session_state[key] = value


def safe_float(value: Any, default: float = 0.0) -> float:
    try:
        x = float(value)
        return x if math.isfinite(x) else default
    except (TypeError, ValueError):
        return default


def calculate_bar(row: dict[str, Any]) -> dict[str, Any]:
    qty = max(0, int(safe_float(row.get("qty"), 1)))
    dia = max(0.0, safe_float(row.get("diameter_mm"), 0))
    a = max(0.0, safe_float(row.get("a_mm"), 0))
    b = max(0.0, safe_float(row.get("b_mm"), 0))
    c = max(0.0, safe_float(row.get("c_mm"), 0))
    d = max(0.0, safe_float(row.get("d_mm"), 0))
    extra = max(0.0, safe_float(row.get("extra_mm"), 0))
    lap = max(0.0, safe_float(row.get("lap_mm"), 0))
    hook = max(0.0, safe_float(row.get("hook_mm"), 0))
    bend_deduction = max(0.0, safe_float(row.get("bend_deduction_mm"), 0))

    shape = str(row.get("shape", "Straight"))
    # Geometry is intentionally transparent: the user supplies dimensions.
    # For bent shapes, a simple leg-sum is used; bend deduction is a user input.
    straight_length = a + b + c + d + extra + lap + hook
    cutting_length = max(0.0, straight_length - bend_deduction)

    unit_weight = dia * dia / 162.0  # kg/m, standard nominal-bar approximation
    total_length_m = cutting_length / 1000.0 * qty
    total_weight_kg = total_length_m * unit_weight

    result = dict(row)
    result.update(
        {
            "qty": qty,
            "diameter_mm": dia,
            "cutting_length_mm": cutting_length,
            "unit_weight_kg_m": unit_weight,
            "total_length_m": total_length_m,
            "total_weight_kg": total_weight_kg,
            "shape": shape,
        }
    )
    return result


def calculate_all(rows: list[dict[str, Any]]) -> pd.DataFrame:
    calculated = [calculate_bar(row) for row in rows]
    columns = [
        "bar_mark",
        "member",
        "bar_type",
        "shape",
        "diameter_mm",
        "spacing_mm",
        "qty",
        "a_mm",
        "b_mm",
        "c_mm",
        "d_mm",
        "extra_mm",
        "lap_mm",
        "hook_mm",
        "bend_deduction_mm",
        "cutting_length_mm",
        "unit_weight_kg_m",
        "total_length_m",
        "total_weight_kg",
        "notes",
    ]
    df = pd.DataFrame(calculated)
    for col in columns:
        if col not in df.columns:
            df[col] = ""
    return df[columns]


def make_excel(df: pd.DataFrame, project_name: str) -> bytes:
    if Workbook is None:
        raise RuntimeError("openpyxl is not installed.")

    wb = Workbook()
    ws = wb.active
    ws.title = "BBS"

    title = f"{project_name} - Bar Bending Schedule"
    ws["A1"] = title
    ws["A1"].font = Font(bold=True, size=16)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max(1, len(df.columns)))

    headers = list(df.columns)
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col_idx, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="D9EAF7")
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for r_idx, values in enumerate(df.itertuples(index=False, name=None), 4):
        for c_idx, value in enumerate(values, 1):
            ws.cell(row=r_idx, column=c_idx, value=value)

    total_row = len(df) + 5
    weight_col = headers.index("total_weight_kg") + 1
    length_col = headers.index("total_length_m") + 1
    ws.cell(row=total_row, column=weight_col, value=f"=SUM({get_column_letter(weight_col)}4:{get_column_letter(weight_col)}{total_row-2})")
    ws.cell(row=total_row, column=length_col, value=f"=SUM({get_column_letter(length_col)}4:{get_column_letter(length_col)}{total_row-2})")
    ws.cell(row=total_row, column=max(1, weight_col - 1), value="TOTAL")

    for column_cells in ws.columns:
        max_len = min(
            45,
            max(len(str(cell.value)) if cell.value is not None else 0 for cell in column_cells) + 2,
        )
        ws.column_dimensions[get_column_letter(column_cells[0].column)].width = max_len

    ws.freeze_panes = "A4"
    ws.auto_filter.ref = ws.dimensions

    summary = wb.create_sheet("Summary")
    summary["A1"] = "BBS Summary"
    summary["A1"].font = Font(bold=True, size=16)
    summary["A3"] = "Project"
    summary["B3"] = project_name
    summary["A4"] = "Total bars"
    summary["B4"] = f"=SUM(BBS!G4:G{len(df)+3})" if not df.empty else 0
    summary["A5"] = "Total length (m)"
    summary["B5"] = f"=SUM(BBS!R4:R{len(df)+3})" if not df.empty else 0
    summary["A6"] = "Total steel (kg)"
    summary["B6"] = f"=SUM(BBS!S4:S{len(df)+3})" if not df.empty else 0
    summary.column_dimensions["A"].width = 25
    summary.column_dimensions["B"].width = 25

    output = io.BytesIO()
    wb.save(output)
    return output.getvalue()


def parse_dxf(uploaded_file) -> tuple[pd.DataFrame, dict[str, Any]]:
    if ezdxf is None:
        raise RuntimeError("ezdxf is not installed.")

    data = uploaded_file.getvalue()
    doc = ezdxf.read(io.BytesIO(data))
    msp = doc.modelspace()

    rows = []
    entity_counts: dict[str, int] = {}
    for entity in msp:
        kind = entity.dxftype()
        entity_counts[kind] = entity_counts.get(kind, 0) + 1

        if kind == "LINE":
            start = entity.dxf.start
            end = entity.dxf.end
            length = math.dist(
                (float(start.x), float(start.y), float(getattr(start, "z", 0))),
                (float(end.x), float(end.y), float(getattr(end, "z", 0))),
            )
            rows.append({"entity": "LINE", "length": length, "layer": entity.dxf.layer})
        elif kind in {"LWPOLYLINE", "POLYLINE"}:
            try:
                length = float(entity.length())
            except Exception:
                length = 0.0
            rows.append({"entity": kind, "length": length, "layer": entity.dxf.layer})
        elif kind == "CIRCLE":
            radius = float(entity.dxf.radius)
            rows.append({"entity": "CIRCLE", "length": 2 * math.pi * radius, "layer": entity.dxf.layer})
        elif kind == "ARC":
            radius = float(entity.dxf.radius)
            angle = math.radians(float(entity.dxf.end_angle) - float(entity.dxf.start_angle))
            if angle < 0:
                angle += 2 * math.pi
            rows.append({"entity": "ARC", "length": radius * angle, "layer": entity.dxf.layer})

    entity_df = pd.DataFrame(rows, columns=["entity", "length", "layer"])
    summary = {
        "version": getattr(doc, "dxfversion", "unknown"),
        "entities": entity_counts,
        "entity_rows": len(entity_df),
    }
    return entity_df, summary


def build_ai_prompt(df: pd.DataFrame, dxf_summary: dict[str, Any] | None, standards: str) -> str:
    records = df.fillna("").to_dict(orient="records")
    payload = {
        "bbs_rows": records,
        "drawing_summary": dxf_summary,
        "requested_review_basis": standards,
    }
    return f"""
You are an experienced reinforcement detailing reviewer.
Review the supplied BBS/drawing-derived data. Identify:
1. missing or suspicious dimensions,
2. duplicate/conflicting bar marks,
3. unusually low/high cutting lengths,
4. questionable quantities or spacing,
5. steel-weight anomalies,
6. constructability/detailing issues,
7. information that must be confirmed from structural drawings.

Do NOT invent structural design requirements. Clearly separate:
- arithmetic/data-quality checks,
- detailing suggestions,
- items requiring a structural engineer's confirmation.

Return a concise professional report with a table-like list of issue, severity, reason, and suggested action.

Review basis: {standards}

DATA:
{json.dumps(payload, indent=2, default=str)}
""".strip()


def run_gemini(prompt: str, model: str) -> str:
    try:
        secret_key = st.secrets.get("GEMINI_API_KEY", "")
    except Exception:
        secret_key = ""
    api_key = secret_key or st.session_state.get("api_key") or os.getenv("GEMINI_API_KEY")
    if not api_key:
        return "Gemini analysis is not configured. Enter your Gemini API key in the sidebar or set GEMINI_API_KEY."

    if genai is None:
        return "The google-genai package is not installed. Add it to requirements.txt and redeploy."

    try:
        client = genai.Client(api_key=api_key)
        response = client.models.generate_content(
            model=model,
            contents=prompt,
        )
        return getattr(response, "text", "") or "Gemini returned no text."
    except Exception as exc:
        return f"Gemini request failed: {type(exc).__name__}: {exc}"


def seed_demo() -> None:
    st.session_state.bars = [
        {
            "bar_mark": "B1",
            "member": "Beam B1",
            "bar_type": "Bottom",
            "shape": "Straight",
            "diameter_mm": 16,
            "spacing_mm": "",
            "qty": 4,
            "a_mm": 5000,
            "b_mm": 0,
            "c_mm": 0,
            "d_mm": 0,
            "extra_mm": 0,
            "lap_mm": 0,
            "hook_mm": 0,
            "bend_deduction_mm": 0,
            "notes": "Demo row",
        },
        {
            "bar_mark": "B2",
            "member": "Beam B1",
            "bar_type": "Stirrup",
            "shape": "Bent",
            "diameter_mm": 8,
            "spacing_mm": 150,
            "qty": 34,
            "a_mm": 850,
            "b_mm": 250,
            "c_mm": 0,
            "d_mm": 0,
            "extra_mm": 0,
            "lap_mm": 0,
            "hook_mm": 160,
            "bend_deduction_mm": 20,
            "notes": "Demo row",
        },
    ]


init_state()

st.title(APP_TITLE)
st.caption(f"Manual BBS + DXF drawing extraction + Gemini review + Excel export • v{APP_VERSION}")

with st.sidebar:
    st.header("⚙️ Project")
    st.session_state.project_name = st.text_input(
        "Project name", value=st.session_state.project_name
    )
    st.session_state.api_key = st.text_input(
        "Gemini API key",
        type="password",
        help="Prefer Streamlit Secrets or GEMINI_API_KEY in production.",
    )
    gemini_model = st.text_input(
        "Gemini Flash model",
        value=os.getenv("GEMINI_MODEL", "gemini-3.7-flash"),
    )
    standards = st.selectbox(
        "Review basis",
        ["General engineering QA", "ASTM/AASHTO references (review only)", "User-defined project standard"],
    )
    if st.button("Load demo BBS"):
        seed_demo()
        st.session_state.bbs_df = calculate_all(st.session_state.bars)
        st.rerun()

    st.info(
        "DWG files are proprietary binary CAD files. This starter version parses "
        "DXF directly. For DWG, convert/export to DXF from AutoCAD or another CAD "
        "tool before analysis. This avoids pretending that a generic Python parser "
        "can reliably interpret every DWG version."
    )

tab1, tab2, tab3, tab4 = st.tabs(
    ["🧮 Manual BBS", "📐 CAD Analysis", "🤖 AI Review", "📊 Export"]
)

with tab1:
    st.subheader("Manual BBS input")
    st.write("Enter one bar mark per row. Dimensions are in mm; steel weight uses d²/162 kg/m.")

    if not st.session_state.bars:
        st.session_state.bars = [
            {
                "bar_mark": "",
                "member": "",
                "bar_type": "Main",
                "shape": "Straight",
                "diameter_mm": 12,
                "spacing_mm": "",
                "qty": 1,
                "a_mm": 0,
                "b_mm": 0,
                "c_mm": 0,
                "d_mm": 0,
                "extra_mm": 0,
                "lap_mm": 0,
                "hook_mm": 0,
                "bend_deduction_mm": 0,
                "notes": "",
            }
        ]

    input_columns = [
        "bar_mark", "member", "bar_type", "shape", "diameter_mm", "spacing_mm",
        "qty", "a_mm", "b_mm", "c_mm", "d_mm", "extra_mm", "lap_mm",
        "hook_mm", "bend_deduction_mm", "notes",
    ]
    edited = st.data_editor(
        pd.DataFrame(st.session_state.bars),
        key="bbs_editor",
        num_rows="dynamic",
        use_container_width=True,
        hide_index=True,
        column_config={
            "bar_type": st.column_config.SelectboxColumn(
                "Bar type", options=["Main", "Top", "Bottom", "Stirrup", "Distribution", "Extra"]
            ),
            "shape": st.column_config.SelectboxColumn(
                "Shape", options=["Straight", "Bent", "L", "U", "Custom"]
            ),
            "diameter_mm": st.column_config.NumberColumn("Dia (mm)", min_value=0, step=1),
            "qty": st.column_config.NumberColumn("Qty", min_value=0, step=1),
        },
    )
    st.session_state.bars = edited.to_dict(orient="records")

    if st.button("Calculate / Update BBS", type="primary", key="calculate_bbs"):
        st.session_state.bbs_df = calculate_all(st.session_state.bars)
        st.success("BBS calculated successfully.")

    if st.session_state.bbs_df is not None:
        df = st.session_state.bbs_df
        st.dataframe(df, use_container_width=True, hide_index=True)
        total_kg = float(pd.to_numeric(df["total_weight_kg"], errors="coerce").fillna(0).sum())
        total_m = float(pd.to_numeric(df["total_length_m"], errors="coerce").fillna(0).sum())
        c1, c2, c3 = st.columns(3)
        c1.metric("Total steel", f"{total_kg:,.2f} kg")
        c2.metric("Total bar length", f"{total_m:,.2f} m")
        c3.metric("Bar marks", f"{len(df):,}")

with tab2:
    st.subheader("CAD drawing analysis")
    cad_file = st.file_uploader(
        "Upload AutoCAD DXF",
        type=["dxf"],
        help="DXF is the reliable CAD interchange format supported by this starter app.",
    )
    if cad_file is not None:
        if st.button("Analyze DXF", type="primary"):
            try:
                entity_df, summary = parse_dxf(cad_file)
                st.session_state.dxf_summary = summary
                st.success(f"DXF read successfully: {summary['entity_rows']} supported geometry rows.")
                st.json(summary)
                st.dataframe(entity_df, use_container_width=True, hide_index=True)

                # Show the longest geometry first as a quick QA aid.
                if not entity_df.empty:
                    st.subheader("Geometry QA")
                    st.dataframe(
                        entity_df.sort_values("length", ascending=False).head(50),
                        use_container_width=True,
                        hide_index=True,
                    )
            except Exception as exc:
                st.error(f"Could not parse DXF: {type(exc).__name__}: {exc}")
    else:
        st.info("Upload a DXF to inspect supported LINE, POLYLINE, LWPOLYLINE, CIRCLE and ARC geometry.")

with tab3:
    st.subheader("Gemini engineering review")
    if st.session_state.bbs_df is None:
        st.warning("Calculate the BBS first.")
    else:
        if st.button("Run Gemini Review", type="primary"):
            prompt = build_ai_prompt(
                st.session_state.bbs_df,
                st.session_state.get("dxf_summary"),
                standards,
            )
            with st.spinner("Gemini is reviewing the data..."):
                st.session_state.analysis = run_gemini(prompt, gemini_model)

        if st.session_state.analysis:
            st.markdown(st.session_state.analysis)

with tab4:
    st.subheader("Excel export")
    if st.session_state.bbs_df is None:
        st.warning("Calculate the BBS first.")
    else:
        df = st.session_state.bbs_df
        try:
            xlsx = make_excel(df, st.session_state.project_name)
            filename = re.sub(r"[^A-Za-z0-9_-]+", "_", st.session_state.project_name).strip("_") or "BBS_Project"
            st.download_button(
                "⬇️ Download BBS Excel",
                data=xlsx,
                file_name=f"{filename}_BBS.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                type="primary",
            )
        except Exception as exc:
            st.error(f"Excel generation failed: {type(exc).__name__}: {exc}")

st.divider()
st.caption(
    "Engineering note: this tool performs data extraction, arithmetic and QA assistance. "
    "It does not replace structural design, approved drawings, or a licensed engineer's review."
)
