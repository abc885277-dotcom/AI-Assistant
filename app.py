import io
import json
import math
import os
import re
from typing import Any

import pandas as pd
import streamlit as st

try:
    import ezdxf
except Exception:
    ezdxf = None

try:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter
except Exception:
    Workbook = None
    Alignment = Font = PatternFill = None
    get_column_letter = None

try:
    from google import genai
except Exception:
    genai = None

st.set_page_config(page_title="AI BBS Engineer", page_icon="🏗️", layout="wide")

APP_TITLE = "🏗️ AI BBS Engineer"
APP_VERSION = "1.1.0"


def init_state() -> None:
    defaults = {
        "bars": [],
        "bbs_df": None,
        "analysis": "",
        "dxf_summary": None,
        "dxf_text": [],
        "project_name": "BBS Project",
        "api_key": "",
    }
    for key, value in defaults.items():
        if key not in st.session_state:
            st.session_state[key] = value


def safe_float(value: Any, default: float = 0.0) -> float:
    try:
        x = float(value)
        return x if math.isfinite(x) else default
    except (TypeError, ValueError):
        return default


def safe_int(value: Any, default: int = 0) -> int:
    try:
        return max(0, int(float(value)))
    except (TypeError, ValueError):
        return default


def kg_per_m(diameter_mm: float) -> float:
    return max(0.0, diameter_mm) ** 2 / 162.0


def development_length_mm(diameter_mm: float, factor: float, minimum_multiplier: float) -> float:
    return max(0.0, diameter_mm) * max(factor, minimum_multiplier)


def normalize_cover_mm(cover: float, member_depth: float) -> float:
    return min(max(cover, 0.0), max(member_depth / 2.0, 0.0))


def automatic_bar_count(clear_distribution_mm: float, spacing_mm: float, minimum: int = 1) -> int:
    """Number of bars from available distribution length and spacing.

    This is a quantity formula, not a structural design check.
    For spacing > 0: floor(clear_length / spacing) + 1, with at least `minimum`.
    """
    if spacing_mm <= 0:
        return minimum
    return max(minimum, int(math.floor(max(clear_distribution_mm, 0.0) / spacing_mm)) + 1)


def calculate_bar(row: dict[str, Any]) -> dict[str, Any]:
    qty = safe_int(row.get("qty"), 1)
    dia = max(0.0, safe_float(row.get("diameter_mm")))
    a = max(0.0, safe_float(row.get("a_mm")))
    b = max(0.0, safe_float(row.get("b_mm")))
    c = max(0.0, safe_float(row.get("c_mm")))
    d = max(0.0, safe_float(row.get("d_mm")))
    extra = max(0.0, safe_float(row.get("extra_mm")))
    lap = max(0.0, safe_float(row.get("lap_mm")))
    hook = max(0.0, safe_float(row.get("hook_mm")))
    bend_deduction = max(0.0, safe_float(row.get("bend_deduction_mm")))

    cutting_length = max(0.0, a + b + c + d + extra + lap + hook - bend_deduction)
    unit_weight = kg_per_m(dia)
    total_length_m = cutting_length / 1000.0 * qty
    total_weight_kg = total_length_m * unit_weight

    result = dict(row)
    result.update(
        {
            "qty": qty,
            "diameter_mm": dia,
            "cutting_length_mm": round(cutting_length, 1),
            "unit_weight_kg_m": round(unit_weight, 3),
            "total_length_m": round(total_length_m, 3),
            "total_weight_kg": round(total_weight_kg, 2),
        }
    )
    return result


def calculate_all(rows: list[dict[str, Any]]) -> pd.DataFrame:
    columns = [
        "bar_mark", "member", "bar_type", "shape", "diameter_mm", "spacing_mm", "qty",
        "a_mm", "b_mm", "c_mm", "d_mm", "extra_mm", "lap_mm", "hook_mm",
        "bend_deduction_mm", "cutting_length_mm", "unit_weight_kg_m", "total_length_m",
        "total_weight_kg", "notes",
    ]
    calculated = [calculate_bar(row) for row in rows]
    df = pd.DataFrame(calculated)
    for col in columns:
        if col not in df.columns:
            df[col] = ""
    return df[columns]


def add_auto_member(member_type: str, member_id: str, length: float, width: float, depth: float,
                    cover: float, dia: float, spacing: float, direction: str, explicit_qty: int,
                    lap: float, hook: float, bend: float, note: str) -> dict[str, Any]:
    effective_cover = normalize_cover_mm(cover, min(depth, width))
    if direction == "Longitudinal":
        bar_length = max(length - 2 * effective_cover, 0.0)
        distribution_length = max(width - 2 * effective_cover, 0.0)
    else:
        bar_length = max(width - 2 * effective_cover, 0.0)
        distribution_length = max(length - 2 * effective_cover, 0.0)

    qty = automatic_bar_count(distribution_length, spacing, 1) if spacing > 0 else max(0, explicit_qty)
    if qty == 0:
        raise ValueError("Enter a positive spacing or an explicit bar quantity.")

    return {
        "bar_mark": f"{member_id}-{direction[:3].upper()}-{int(dia)}",
        "member": member_id,
        "bar_type": "Main",
        "shape": "Straight",
        "diameter_mm": dia,
        "spacing_mm": spacing if spacing > 0 else "",
        "qty": qty,
        "a_mm": bar_length,
        "b_mm": 0,
        "c_mm": 0,
        "d_mm": 0,
        "extra_mm": 0,
        "lap_mm": lap,
        "hook_mm": hook,
        "bend_deduction_mm": bend,
        "notes": note,
    }


def make_excel(df: pd.DataFrame, project_name: str) -> bytes:
    if Workbook is None:
        raise RuntimeError("openpyxl is not installed. Add openpyxl to requirements.txt.")

    wb = Workbook()
    ws = wb.active
    ws.title = "BBS"
    title = f"{project_name} - Bar Bending Schedule"
    ws["A1"] = title
    ws["A1"].font = Font(bold=True, size=16)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max(1, len(df.columns)))

    headers = list(df.columns)
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col_idx, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1F4E78")
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for r_idx, values in enumerate(df.itertuples(index=False, name=None), 4):
        for c_idx, value in enumerate(values, 1):
            ws.cell(row=r_idx, column=c_idx, value=value)

    data_end = len(df) + 3
    total_row = data_end + 1
    if df.empty:
        ws.cell(total_row, 1, "TOTAL")
    else:
        length_col = headers.index("total_length_m") + 1
        weight_col = headers.index("total_weight_kg") + 1
        ws.cell(total_row, max(1, weight_col - 1), "TOTAL")
        ws.cell(total_row, length_col, f"=SUM({get_column_letter(length_col)}4:{get_column_letter(length_col)}{data_end})")
        ws.cell(total_row, weight_col, f"=SUM({get_column_letter(weight_col)}4:{get_column_letter(weight_col)}{data_end})")
        ws.cell(total_row, length_col).font = Font(bold=True)
        ws.cell(total_row, weight_col).font = Font(bold=True)

    for column_cells in ws.columns:
        max_len = min(45, max(len(str(c.value)) if c.value is not None else 0 for c in column_cells) + 2)
        ws.column_dimensions[get_column_letter(column_cells[0].column)].width = max(10, max_len)
    ws.freeze_panes = "A4"
    if not df.empty:
        ws.auto_filter.ref = f"A3:{get_column_letter(len(headers))}{data_end}"

    summary = wb.create_sheet("Summary")
    summary["A1"] = "BBS Summary"
    summary["A1"].font = Font(bold=True, size=16)
    summary["A3"] = "Project"
    summary["B3"] = project_name
    summary["A4"] = "BBS rows"
    summary["B4"] = len(df)
    summary["A5"] = "Total quantity"
    summary["B5"] = int(pd.to_numeric(df.get("qty", pd.Series(dtype=float)), errors="coerce").fillna(0).sum())
    summary["A6"] = "Total length (m)"
    summary["B6"] = float(pd.to_numeric(df.get("total_length_m", pd.Series(dtype=float)), errors="coerce").fillna(0).sum())
    summary["A7"] = "Total steel (kg)"
    summary["B7"] = float(pd.to_numeric(df.get("total_weight_kg", pd.Series(dtype=float)), errors="coerce").fillna(0).sum())
    summary.column_dimensions["A"].width = 25
    summary.column_dimensions["B"].width = 25

    output = io.BytesIO()
    wb.save(output)
    return output.getvalue()


def parse_rebar_text(text: str) -> list[dict[str, Any]]:
    """Find common reinforcement callouts in DXF text without claiming full drawing interpretation."""
    found = []
    normalized = re.sub(r"\s+", " ", text or "").strip()
    patterns = [
        re.compile(r"(?P<qty>\d+)\s*[ØΦφ]\s*(?P<dia>\d+(?:\.\d+)?)", re.I),
        re.compile(r"(?P<qty>\d+)\s*[YyTt]\s*(?P<dia>\d+(?:\.\d+)?)", re.I),
        re.compile(r"[ØΦφ]\s*(?P<dia>\d+(?:\.\d+)?)\s*[@]\s*(?P<spacing>\d+(?:\.\d+)?)", re.I),
        re.compile(r"(?P<dia>\d+(?:\.\d+)?)\s*[@]\s*(?P<spacing>\d+(?:\.\d+)?)", re.I),
    ]
    for pattern in patterns:
        for match in pattern.finditer(normalized):
            item = match.groupdict()
            item["raw"] = match.group(0)
            found.append(item)
    return found


def parse_dxf(uploaded_file) -> tuple[pd.DataFrame, dict[str, Any]]:
    if ezdxf is None:
        raise RuntimeError("ezdxf is not installed. Add ezdxf to requirements.txt.")
    data = uploaded_file.getvalue()
    doc = ezdxf.read(io.BytesIO(data))
    msp = doc.modelspace()
    rows = []
    entity_counts: dict[str, int] = {}
    text_items: list[str] = []

    for entity in msp:
        kind = entity.dxftype()
        entity_counts[kind] = entity_counts.get(kind, 0) + 1
        layer = getattr(entity.dxf, "layer", "")
        if kind == "LINE":
            start, end = entity.dxf.start, entity.dxf.end
            length = math.dist((float(start.x), float(start.y), float(getattr(start, "z", 0))),
                               (float(end.x), float(end.y), float(getattr(end, "z", 0))))
            rows.append({"entity": kind, "length": length, "layer": layer, "text": ""})
        elif kind in {"LWPOLYLINE", "POLYLINE"}:
            try:
                length = float(entity.length())
            except Exception:
                length = 0.0
            rows.append({"entity": kind, "length": length, "layer": layer, "text": ""})
        elif kind == "CIRCLE":
            radius = float(entity.dxf.radius)
            rows.append({"entity": kind, "length": 2 * math.pi * radius, "layer": layer, "text": ""})
        elif kind == "ARC":
            radius = float(entity.dxf.radius)
            angle = math.radians(float(entity.dxf.end_angle) - float(entity.dxf.start_angle))
            if angle < 0:
                angle += 2 * math.pi
            rows.append({"entity": kind, "length": radius * angle, "layer": layer, "text": ""})
        elif kind == "TEXT":
            value = str(getattr(entity.dxf, "text", ""))
            text_items.append(value)
            rows.append({"entity": kind, "length": 0.0, "layer": layer, "text": value})
        elif kind == "MTEXT":
            try:
                value = str(entity.text)
            except Exception:
                value = str(getattr(entity.dxf, "text", ""))
            text_items.append(value)
            rows.append({"entity": kind, "length": 0.0, "layer": layer, "text": value})

    entity_df = pd.DataFrame(rows, columns=["entity", "length", "layer", "text"])
    text_blob = "\n".join(text_items)
    rebar_matches = parse_rebar_text(text_blob)
    summary = {
        "version": getattr(doc, "dxfversion", "unknown"),
        "entities": entity_counts,
        "entity_rows": len(entity_df),
        "text_count": len(text_items),
        "reinforcement_callouts_detected": len(rebar_matches),
        "reinforcement_matches": rebar_matches,
    }
    return entity_df, summary


def build_ai_prompt(df: pd.DataFrame, dxf_summary: dict[str, Any] | None, standards: str) -> str:
    payload = {
        "bbs_rows": df.fillna("").to_dict(orient="records"),
        "drawing_summary": dxf_summary,
        "requested_review_basis": standards,
    }
    return f"""You are an experienced reinforcement detailing reviewer.\n\nReview the supplied BBS/drawing-derived data. Identify:\n1. missing or suspicious dimensions,\n2. duplicate/conflicting bar marks,\n3. unusual cutting lengths,\n4. questionable quantities or spacing,\n5. steel-weight anomalies,\n6. constructability/detailing questions,\n7. information that must be confirmed from structural drawings.\n\nDo NOT certify structural adequacy and do NOT invent missing design requirements. Separate arithmetic/data-quality checks from detailing questions and engineer/code verification items.\n\nReview basis: {standards}\n\nDATA:\n{json.dumps(payload, indent=2, default=str)}"""


def run_gemini(prompt: str, model: str) -> str:
    try:
        secret_key = st.secrets.get("GEMINI_API_KEY", "")
    except Exception:
        secret_key = ""
    api_key = secret_key or st.session_state.get("api_key") or os.getenv("GEMINI_API_KEY")
    if not api_key:
        return "Gemini is not configured. Enter the API key in the sidebar or set GEMINI_API_KEY in Streamlit Secrets."
    if genai is None:
        return "google-genai is not installed. Add it to requirements.txt."
    try:
        client = genai.Client(api_key=api_key)
        response = client.models.generate_content(model=model, contents=prompt)
        return getattr(response, "text", "") or "Gemini returned no text."
    except Exception as exc:
        return f"Gemini request failed: {type(exc).__name__}: {exc}"


def seed_demo() -> None:
    st.session_state.bars = [
        {"bar_mark":"B1", "member":"Beam B1", "bar_type":"Bottom", "shape":"Straight", "diameter_mm":16,
         "spacing_mm":"", "qty":4, "a_mm":5000, "b_mm":0, "c_mm":0, "d_mm":0, "extra_mm":0, "lap_mm":0,
         "hook_mm":0, "bend_deduction_mm":0, "notes":"Demo row"},
        {"bar_mark":"B2", "member":"Beam B1", "bar_type":"Stirrup", "shape":"Bent", "diameter_mm":8,
         "spacing_mm":150, "qty":34, "a_mm":850, "b_mm":250, "c_mm":0, "d_mm":0, "extra_mm":0, "lap_mm":0,
         "hook_mm":160, "bend_deduction_mm":20, "notes":"Demo row"},
    ]
    st.session_state.bbs_df = calculate_all(st.session_state.bars)


init_state()

st.title(APP_TITLE)
st.caption(f"Automatic BBS calculation + DXF extraction + Gemini review + Excel export • v{APP_VERSION}")

with st.sidebar:
    st.header("⚙️ Project")
    st.session_state.project_name = st.text_input("Project name", st.session_state.project_name)
    st.session_state.api_key = st.text_input("Gemini API key", type="password")
    gemini_model = st.text_input("Gemini Flash model", os.getenv("GEMINI_MODEL", "gemini-3.7-flash"))
    standards = st.selectbox("Review basis", ["General engineering QA", "ASTM/AASHTO references (review only)", "User-defined project standard"])
    if st.button("Load demo BBS"):
        seed_demo()
        st.rerun()
    st.info("DXF is supported directly. DWG should be exported/converted to DXF before parsing.")

# Shared assumptions
st.subheader("BBS assumptions")
a1, a2, a3, a4, a5 = st.columns(5)
with a1: cover = st.number_input("Cover (mm)", 0.0, 200.0, 40.0, 1.0)
with a2: dev_factor = st.number_input("DL factor × dia", 1.0, 100.0, 40.0, 1.0)
with a3: dev_min_mult = st.number_input("Minimum DL × dia", 1.0, 100.0, 12.0, 1.0)
with a4: hook = st.number_input("Hook/end (mm)", 0.0, 1000.0, 0.0, 10.0)
with a5: bend = st.number_input("Bend deduction (mm)", 0.0, 3000.0, 0.0, 10.0)
lap = st.number_input("Lap addition per bar (mm)", 0.0, 5000.0, 0.0, 10.0)

st.warning("The calculation engine performs quantity/arithmetic assistance. Development length, hooks, bends, laps and final detailing must be verified against the governing code and approved structural drawings.")

tab1, tab2, tab3, tab4 = st.tabs(["🧮 Manual / Auto BBS", "📐 CAD Analysis", "🤖 AI Review", "📊 Excel Export"])

with tab1:
    st.subheader("Automatic member calculation")
    c1, c2, c3, c4 = st.columns(4)
    with c1: member_type = st.selectbox("Member type", ["Slab", "Beam", "Column"])
    with c2: member_id = st.text_input("Member ID", f"{member_type[:1].upper()}-01")
    with c3: length = st.number_input("Length (mm)", 100.0, 100000.0, 5000.0, 50.0)
    with c4: width = st.number_input("Width (mm)", 100.0, 100000.0, 3000.0, 50.0)
    c5, c6, c7, c8 = st.columns(4)
    with c5: depth = st.number_input("Depth / thickness (mm)", 50.0, 5000.0, 150.0, 10.0)
    with c6: dia = st.selectbox("Bar diameter (mm)", [6,8,10,12,16,20,25,32,40], index=3)
    with c7: spacing = st.number_input("Spacing (mm)", 0.0, 2000.0, 150.0, 10.0)
    with c8: direction = st.selectbox("Direction", ["Longitudinal", "Transverse"])
    explicit_qty = st.number_input("Explicit quantity (used when spacing = 0)", 0, 10000, 0, 1)
    note = st.text_input("Detailing note", "Verify against structural drawings.")

    if st.button("➕ Add Automatically Calculated Bar", type="primary", use_container_width=True):
        try:
            row = add_auto_member(member_type, member_id, length, width, depth, cover, float(dia), spacing, direction,
                                  int(explicit_qty), lap, hook, bend, note)
            st.session_state.bars.append(row)
            st.session_state.bbs_df = calculate_all(st.session_state.bars)
            st.success(f"Added {member_id}. Bars and steel weight calculated automatically.")
        except ValueError as exc:
            st.error(str(exc))

    st.subheader("BBS input table")
    if not st.session_state.bars:
        st.session_state.bars = [{"bar_mark":"", "member":"", "bar_type":"Main", "shape":"Straight", "diameter_mm":12,
                                  "spacing_mm":"", "qty":1, "a_mm":0, "b_mm":0, "c_mm":0, "d_mm":0, "extra_mm":0,
                                  "lap_mm":0, "hook_mm":0, "bend_deduction_mm":0, "notes":""}]

    edited = st.data_editor(
        pd.DataFrame(st.session_state.bars),
        key="bbs_editor",
        num_rows="dynamic",
        use_container_width=True,
        hide_index=True,
        column_config={
            "bar_type": st.column_config.SelectboxColumn("Bar type", options=["Main","Top","Bottom","Stirrup","Distribution","Extra"]),
            "shape": st.column_config.SelectboxColumn("Shape", options=["Straight","Bent","L","U","Custom"]),
            "diameter_mm": st.column_config.NumberColumn("Dia (mm)", min_value=0, step=1),
            "qty": st.column_config.NumberColumn("Qty", min_value=0, step=1),
            "spacing_mm": st.column_config.NumberColumn("Spacing (mm)", min_value=0, step=1),
        },
    )
    st.session_state.bars = edited.to_dict(orient="records")

    # Automatic recalculation on every Streamlit rerun; no separate calculate button required.
    st.session_state.bbs_df = calculate_all(st.session_state.bars)
    df = st.session_state.bbs_df
    st.dataframe(df, use_container_width=True, hide_index=True)

    total_kg = float(pd.to_numeric(df["total_weight_kg"], errors="coerce").fillna(0).sum()) if not df.empty else 0.0
    total_m = float(pd.to_numeric(df["total_length_m"], errors="coerce").fillna(0).sum()) if not df.empty else 0.0
    total_qty = int(pd.to_numeric(df["qty"], errors="coerce").fillna(0).sum()) if not df.empty else 0
    m1, m2, m3 = st.columns(3)
    m1.metric("Total quantity", f"{total_qty:,}")
    m2.metric("Total length", f"{total_m:,.2f} m")
    m3.metric("Total steel", f"{total_kg:,.2f} kg")

    if not df.empty:
        summary_df = df.groupby("diameter_mm", as_index=False)["total_weight_kg"].sum().rename(columns={"diameter_mm":"Dia (mm)", "total_weight_kg":"Steel Weight (kg)"})
        summary_df["Steel Weight (kg)"] = summary_df["Steel Weight (kg)"].round(2)
        st.subheader("Steel summary by diameter")
        st.dataframe(summary_df, use_container_width=True, hide_index=True)
    else:
        summary_df = pd.DataFrame(columns=["Dia (mm)", "Steel Weight (kg)"])

    if st.button("🗑️ Clear all BBS rows"):
        st.session_state.bars = []
        st.session_state.bbs_df = calculate_all([])
        st.rerun()

with tab2:
    st.subheader("DXF drawing analysis")
    cad_file = st.file_uploader("Upload AutoCAD DXF", type=["dxf"], help="Direct DXF parsing with geometry and text extraction.")
    if cad_file is not None and st.button("Analyze DXF", type="primary"):
        try:
            entity_df, summary = parse_dxf(cad_file)
            st.session_state.dxf_summary = summary
            st.success(f"DXF read successfully: {summary['entity_rows']} entity rows.")
            st.json(summary)
            st.dataframe(entity_df, use_container_width=True, hide_index=True)
            if summary["reinforcement_matches"]:
                st.subheader("Detected reinforcement callouts")
                st.dataframe(pd.DataFrame(summary["reinforcement_matches"]), use_container_width=True, hide_index=True)
            if not entity_df.empty:
                st.subheader("Geometry QA")
                st.dataframe(entity_df.sort_values("length", ascending=False).head(50), use_container_width=True, hide_index=True)
        except Exception as exc:
            st.error(f"Could not parse DXF: {type(exc).__name__}: {exc}")
    else:
        st.info("Upload a DXF to inspect LINE, POLYLINE, LWPOLYLINE, CIRCLE, ARC, TEXT and MTEXT entities.")

with tab3:
    st.subheader("Gemini Flash engineering review")
    if st.session_state.bbs_df is None or st.session_state.bbs_df.empty:
        st.warning("Add or calculate at least one BBS row first.")
    elif st.button("Run Gemini Review", type="primary"):
        prompt = build_ai_prompt(st.session_state.bbs_df, st.session_state.get("dxf_summary"), standards)
        with st.spinner("Gemini is reviewing the data..."):
            st.session_state.analysis = run_gemini(prompt, gemini_model)
    if st.session_state.analysis:
        st.markdown(st.session_state.analysis)

with tab4:
    st.subheader("Excel export")
    df = st.session_state.bbs_df if st.session_state.bbs_df is not None else pd.DataFrame()
    if df.empty:
        st.warning("Add at least one BBS row first.")
    else:
        try:
            xlsx = make_excel(df, st.session_state.project_name)
            filename = re.sub(r"[^A-Za-z0-9_-]+", "_", st.session_state.project_name).strip("_") or "BBS_Project"
            st.download_button("⬇️ Download BBS Excel", data=xlsx, file_name=f"{filename}_BBS.xlsx",
                               mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", type="primary")
        except Exception as exc:
            st.error(f"Excel generation failed: {type(exc).__name__}: {exc}")

st.divider()
st.caption("Engineering note: this application performs deterministic quantity calculations and QA assistance. It does not replace structural design, approved drawings, or licensed engineering review.")
